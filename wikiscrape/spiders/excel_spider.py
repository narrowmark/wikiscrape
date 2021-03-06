import scrapy
import re
import openpyxl
import logging

class ExcelOutputSpider(scrapy.Spider):
  name = "excel"
  start_urls = [
    "https://en.wikipedia.org/wiki/Portal:Current_events",
  ]

  #logging.basicConfig(filename="excel.log", level=logging.INFO)

  def open_or_create_sheet(self):
    # Test for file existence.
    try:
      self.wb = openpyxl.load_workbook('events.xlsx')
    except IOError:
      self.wb = openpyxl.Workbook()

    # Currently changes sheet title whether events.xlsx exists or not.
    # Move around later.
    self.sheet = self.wb.active
    self.sheet.title = "CurrentEvents"
    self.wb.save('events.xlsx')

  def date_parse(self, date):
    self.open_or_create_sheet()

    # date_list looks like:
    # [u'Current events of',
    #  u' January\x04,\xa02017',
    #  u'\xa0(',
    #  u'2017-01.04',
    #  u')',
    #  u' (Wedesnday)']
    date_list = date.xpath(".//text()").extract()

    # Extranous elements from date_list removed.
    # Date converted from Unicode. Will consider changing later pending
    # determination of how Excel handles Unicode.
    date_list = re.findall(r'(?u)\w+', date_list[1])
    date_list = [x.encode('utf-8') for x in date_list]
    month, day, year = date_list[0], date_list[1], date_list[2]

    return month + " " + day + ", " + year

  def headline_parse(self, event):
    # Query for removing external links
    not_external = ".//a[not(contains(@class, 'external text'))]"

    # Only return ongoing events
    if re.findall(r'<li><a.*</a>\n', event.extract()) == []:
      return []

    # head of the form
    # United Nations Security Council election, 2016
    head = event.xpath(not_external + "/text()|text()")[0].extract()
    # event of the form
    # [u'UNited Nations Security Council election, 2016', u'Bolivia', u', ' ...]
    event = event.xpath(".//li/text()|" + not_external +
                        "/text()").extract()

    body = ""
    # event starts with head, which this code omits in adding to body
    for line in event[1:]:
      body += line

    return [head, body]

  def parse(self, response):
    summaries = response.xpath("//table[@class='vevent']")
    # dict of the form { headline: [most_recent_entry, earliest_entry,
    # number_of_entries_about_headline] }
    # Rudimentary form of what will ultimately become a row in an Excel
    # spreadsheet, an entry in a JSON file, etc.
    current = {}

    for summary in summaries:
      # date is a Selector
      # Consists of several components buried inside of different tags that need
      # to be parsed.
      date = summary.xpath(".//span[@class='summary']")[0]
      date = self.date_parse(date)

      descriptions = summary.xpath("descendant::td[@class='description']")
      # desc is a box around all of the events that occur inside of a single day
      for desc in descriptions:
        # headline is a single news item, e.g. a trail derailment
        for headline in desc.xpath("child::ul/li"):
          parsed = self.headline_parse(headline)
          if parsed == []:
            continue

          #logging.info("HEAD: " + parsed[0])
          #logging.info("BODY: " + parsed[1] + "\n")

          if current.has_key(parsed[0]) == False:
            current[parsed[0]] = [date, date, 1]
          else:
            current[parsed[0]][1] = date
            current[parsed[0]][2] += 1

      #logging.info("CURRENT: " + str(current))

      # Populating the spreadsheet
      for e, i in zip(current, range(len(current))):
        self.sheet.cell(row=i+1, column=1).value = e
        self.sheet.cell(row=i+1, column=2).value = current[e][0]
        self.sheet.cell(row=i+1, column=3).value = current[e][1]
        self.sheet.cell(row=i+1, column=4).value = current[e][2]
      self.wb.save('events.xlsx')
